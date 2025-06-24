from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Define the target directory for spreadsheets
BASE_DIR = Path(__file__).resolve().parents[1] / "company_spreadsheets"
BASE_DIR.mkdir(exist_ok=True)

# Spreadsheet filename
filename = "design_bending_capacity_formula_timber.xlsx"
file_path = BASE_DIR / filename

# Only create the file if it doesn't already exist to avoid overwriting user data
if file_path.exists():
    print(f"{filename} already exists – no action taken")
    raise SystemExit(0)

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Timber Bending Capacity"

# Header
ws["A1"] = "Variable"
ws["B1"] = "Value"
ws["C1"] = "Definition"

# Data rows (variable, definition)
rows = [
    ("phi", "Capacity factor (Clause 2.3)"),
    ("k1", "Duration of load factor (Clause 2.4.1.1)"),
    ("k4", "Partial seasoning factor (Clause 2.4.2)"),
    ("k6", "Temperature factor (Clause 2.4.3)"),
    ("k9", "Strength sharing factor (member type clause)"),
    ("k12", "Stability factor (lateral buckling)"),
    ("fb'", "Characteristic bending strength, MPa"),
    ("Z", "Section modulus of beam, mm^3"),
]

start_row = 4  # leave some spacing
for idx, (var, definition) in enumerate(rows, start=start_row):
    ws.cell(row=idx, column=1, value=var)
    ws.cell(row=idx, column=3, value=definition)

# Result row for Md
result_row = start_row + len(rows) + 2  # gap of 1 row
ws.cell(row=result_row, column=1, value="Md")

# Construct formula referencing the value column (B)
value_cells = [f"B{start_row + i}" for i in range(len(rows))]
formula = "=" + "*".join(value_cells)
ws.cell(row=result_row, column=2, value=formula)
ws.cell(row=result_row, column=3, value="Design bending moment capacity")

# Adjust column widths for readability
for col_idx in range(1, 4):
    ws.column_dimensions[get_column_letter(col_idx)].width = 22

# Save workbook
wb.save(file_path)
print(f"Created {file_path}")
