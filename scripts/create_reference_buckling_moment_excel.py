from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Directory for spreadsheets
BASE_DIR = Path(__file__).resolve().parents[1] / "company_spreadsheets"
BASE_DIR.mkdir(exist_ok=True)

filename = "reference_buckling_moment.xlsx"
file_path = BASE_DIR / filename

if file_path.exists():
    print(f"{filename} already exists – no changes made.")
    raise SystemExit(0)

wb = Workbook()
ws = wb.active
ws.title = "Reference Buckling Moment"

# Headers
ws["A1"] = "Variable"
ws["B1"] = "Value"
ws["C1"] = "Definition / Units"

rows = [
    ("E", "Elastic modulus, MPa"),
    ("G", "Shear modulus, MPa"),
    ("I_y", "Second moment of area about y-axis, mm^4"),
    ("J", "Torsion constant, mm^4"),
    ("I_w", "Warping constant, mm^6"),
    ("l_e", "Effective length, mm"),
]
start_row = 4
for idx, (var, definition) in enumerate(rows, start=start_row):
    ws.cell(row=idx, column=1, value=var)
    ws.cell(row=idx, column=3, value=definition)

# Output row
output_row = start_row + len(rows) + 2
ws.cell(row=output_row, column=1, value="M_o")

# Build Excel formula using cell references
# Map: B4=E, B5=G, B6=I_y, B7=J, B8=I_w, B9=l_e
formula = "=SQRT((PI()^2*B4*B6/B9^2)*(B5*B7+PI()^2*B4*B8/B9^2))"
ws.cell(row=output_row, column=2, value=formula)
ws.cell(row=output_row, column=3, value="Reference buckling moment, N·mm")

# Adjust column widths
for col in range(1, 4):
    ws.column_dimensions[get_column_letter(col)].width = 28

wb.save(file_path)
print(f"Created {file_path}")
