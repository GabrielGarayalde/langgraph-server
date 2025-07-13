"""This module provides example tools for web scraping and search functionality.

It includes a basic Tavily search function (as an example)

These tools are intended as free examples to get started. For production use,
consider implementing more robust and specialized tools tailored to your needs.
"""

import os
import json
from pathlib import Path
from typing import Annotated, Any, Callable, Dict, List, Optional, Union, cast
import base64
import asyncio
from io import BytesIO

from dotenv import load_dotenv
from langchain_core.tools import tool
from langchain_tavily import TavilySearch  # type: ignore[import-not-found]

from react_agent.configuration import Configuration

# Pinecone imports and config
try:
    from pinecone import Pinecone
except ImportError:
    Pinecone = None

load_dotenv()
PINECONE_API_KEY = os.getenv("PINECONE_API_KEY")

# Medical Pinecone Configuration
MEDICAL_PINECONE_INDEX_NAME = "anna-medical"
MEDICAL_DEFAULT_NAMESPACE = "anna-medical-namespace"

# Engineering Pinecone Configuration
ENGINEERING_PINECONE_INDEX_NAME = "engineering-index"
ENGINEERING_DEFAULT_NAMESPACE = "engineering-namespace"
# Engineering sparse index name for lexical retrieval
ENGINEERING_SPARSE_PINECONE_INDEX_NAME = "engineering-index-sparse"

_pinecone_client = None

def get_pinecone_client():
    """Get or initialize the Pinecone client instance.
    
    Returns:
        Pinecone client instance or None if not configured.
    """
    global _pinecone_client
    if _pinecone_client is None and PINECONE_API_KEY and Pinecone:
        _pinecone_client = Pinecone(api_key=PINECONE_API_KEY)
    return _pinecone_client


async def search(query: str) -> Optional[dict[str, Any]]:
    """Search for general web results.

    This function performs a search using the Tavily search engine, which is designed
    to provide comprehensive, accurate, and trusted results. It's particularly useful
    for answering questions about current events.
    """
    configuration = Configuration.from_context()
    wrapped = TavilySearch(max_results=configuration.max_search_results)
    return cast(dict[str, Any], await wrapped.ainvoke({"query": query}))


async def calculator(expression: str) -> float:
    """Evaluate a simple arithmetic expression.

    Args:
        expression: A string representing a simple arithmetic expression
            (e.g., "2 + 2", "10 * 5", "100 / 4").
            Supports addition (+), subtraction (-), multiplication (*), and division (/).

    Returns:
        The result of the calculation.
    """
    try:
        # A simple and relatively safe way to evaluate basic math expressions.
        # For production, consider a more robust math expression parser.
        allowed_chars = "0123456789+-*/(). "
        if not all(char in allowed_chars for char in expression):
            raise ValueError("Expression contains invalid characters.")
        # pylint: disable=eval-used
        result = eval(expression)
        return float(result)
    except Exception as e:
        return f"Error evaluating expression '{expression}': {e}"


@tool
def search_engineering_database(
    query: Annotated[str, "Search query for engineering documents and standards"],
    top_k: Annotated[int, "Number of results to return (default: 10)"] = 10,
    namespace: Annotated[str, "Database namespace to search (default: aus-standards-namespace)"] = ENGINEERING_DEFAULT_NAMESPACE,
    source_document_id: Annotated[Optional[str], "Optional filter – only return chunks from this document ID"] = None,
) -> Dict[str, Any]:
    """Search the Pinecone vector database for engineering documents and standards.
    
    Returns a structured dictionary containing the search results, including metadata and content for each hit.
    This tool searches through Australian Standards related to engineering.

    If *source_document_id* is supplied, the search is restricted to that document via a Pinecone metadata
    filter.
    
    Use this tool to find:
    - Specific clauses and sections from australian standards
    - Design rules and formulas for steel structures
    - Material properties and specifications
    - Commentary and explanations related to the standard
    """
    # Check if Pinecone is configured
    if not PINECONE_API_KEY:
        return {
            "type": "database_search_error",
            "error": "PINECONE_API_KEY not found in environment variables. Please configure Pinecone access."
        }
    try:
        # Get Pinecone client
        pc = get_pinecone_client()
        if not pc:
            return {
                "type": "database_search_error",
                "error": "Failed to initialize Pinecone client."
            }
        # Get both dense and sparse Pinecone indexes
        index_dense = pc.Index(ENGINEERING_PINECONE_INDEX_NAME)
        index_sparse = pc.Index(ENGINEERING_SPARSE_PINECONE_INDEX_NAME)

        # Build the query dict with optional metadata filter
        query_dict: Dict[str, Any] = {
            "top_k": min(top_k, 10),
            "inputs": {"text": query},
        }
        if source_document_id:
            # Apply metadata filter
            query_dict["filter"] = {"source_document_id": source_document_id}

        # Common search parameters for both searches
        search_params = {
            "namespace": namespace,
            "query": query_dict,
            "rerank": {
                "model": "bge-reranker-v2-m3",
                "top_n": min(top_k, 10),
                "rank_fields": ["chunk_text"],
            },
        }

        # Execute searches on both indexes
        dense_response = index_dense.search(**search_params)
        sparse_response = index_sparse.search(**search_params)

        # Combine and deduplicate hits
        combined_hits = dense_response.get("result", {}).get("hits", []) + \
                        sparse_response.get("result", {}).get("hits", [])

        # If no hits at all, return early (handled later in existing logic)
        if not combined_hits:
            pinecone_response = {"result": {"hits": []}}
        else:
            seen_keys: set[tuple[str, str]] = set()
            deduped_hits = []
            for hit in combined_hits:
                fields = hit.get("fields", {})
                # Uniqueness key: document + page number (falls back to chunk text hash)
                key = (
                    fields.get("source_document_id", ""),
                    str(fields.get("page_number", ""))
                )
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                deduped_hits.append(hit)

            # Sort by descending Pinecone score and truncate to top_k
            deduped_hits.sort(key=lambda h: h.get("_score", 0.0), reverse=True)
            deduped_hits = deduped_hits[: min(top_k, 10)]
            pinecone_response = {"result": {"hits": deduped_hits}}
        # Process results
        if not pinecone_response.get("result", {}).get("hits"):
            return {
                "type": "database_search_results",
                "query": query,
                "message": f"No relevant documents found for query: '{query}'",
                "results": []
            }
        # Format the results
        processed_results = []
        for hit in pinecone_response["result"]["hits"]:
            fields = hit.get("fields", {})
            result_item = {
                "source_document_id": fields.get("source_document_id", "Unknown Document"),
                "page_number": fields.get("page_number", "Unknown Page"),
                "score": hit.get("_score", 0.0),
                "content": fields.get("chunk_text", ""),
                "clauses_mentioned": fields.get("clauses_mentioned"),
                "tables_mentioned": fields.get("tables_mentioned"),
                "figures_mentioned": fields.get("figures_mentioned")
            }
            processed_results.append(result_item)
        return {
            "type": "database_search_results",
            "query": query,
            "count": len(processed_results),
            "results": processed_results
        }
    except Exception as e:
        return {
            "type": "database_search_error",
            "error": f"Error searching engineering database: {str(e)}",
            "details": "Please check your Pinecone configuration and try again."
        }


@tool
def search_engineering_database_filtered(
    query: Annotated[str, "Search query for engineering documents (use empty string '' for metadata-only filtering)"],
    source_document_id: Annotated[Optional[str], "Filter by specific document ID (e.g., 'as_1720.1_2010')"] = None,
    page_number: Annotated[Optional[int], "Filter by specific page number"] = None,
    tables_mentioned: Annotated[Optional[List[str]], "Filter by tables mentioned in chunks (e.g., ['Table 5.1', 'Table 5.2'])"] = None,
    figures_mentioned: Annotated[Optional[List[str]], "Filter by figures mentioned in chunks (e.g., ['FIGURE 5.3'])"] = None,
    clauses_mentioned: Annotated[Optional[List[str]], "Filter by clauses mentioned in chunks (e.g., ['Clause 2.3'])"] = None,
    top_k: Annotated[int, "Number of results to return (default: 20 for filtered searches)"] = 20,
    namespace: Annotated[str, "Database namespace to search"] = ENGINEERING_DEFAULT_NAMESPACE
) -> Dict[str, Any]:
    """Search the engineering database with metadata filters and return concatenated page content.
    
    This tool retrieves chunks based on metadata filters and concatenates them in sequential order.
    It's particularly useful for:
    - Getting the full text content from a specific page of a document
    - Finding all content that mentions specific tables or figures
    - Retrieving complete context around formulas, tables, or complex content
    
    The query parameter can be empty string '' if you only want to filter by metadata.
    Results are returned as concatenated text in natural page order (by page number and chunk index).
    
    Example use cases:
    1. Get all content from page 90 of AS 1720.1-2010:
       source_document_id='as_1720.1_2010', page_number=90, query=''
    
    2. Find all content mentioning Table 5.1:
       tables_mentioned=['Table 5.1'], query=''
    
    3. Get content from a specific page that mentions certain clauses:
       source_document_id='as_4100_1998', page_number=45, clauses_mentioned=['Clause 3.2'], query=''
    
    Returns:
        A dictionary containing:
        - concatenated_content: All chunk text concatenated in sequential order
        - chunk_count: Number of chunks found
        - pages_covered: List of page numbers included
        - all_clauses_mentioned: All unique clauses mentioned across chunks
        - all_tables_mentioned: All unique tables mentioned across chunks
        - all_figures_mentioned: All unique figures mentioned across chunks
    """
    # Check if Pinecone is configured
    if not PINECONE_API_KEY:
        return {
            "type": "database_search_error",
            "error": "PINECONE_API_KEY not found in environment variables. Please configure Pinecone access."
        }
    
    try:
        # Get Pinecone client
        pc = get_pinecone_client()
        if not pc:
            return {
                "type": "database_search_error",
                "error": "Failed to initialize Pinecone client."
            }
        
        # Get the dense index (we'll use dense for filtered searches)
        index = pc.Index(ENGINEERING_PINECONE_INDEX_NAME)
        
        # Build the filter expression
        filter_conditions = []
        
        if source_document_id:
            filter_conditions.append({"source_document_id": {"$eq": source_document_id}})
        
        if page_number is not None:
            filter_conditions.append({"page_number": {"$eq": page_number}})
        
        if tables_mentioned:
            # Filter for chunks that mention ANY of the specified tables
            table_conditions = [{"tables_mentioned": {"$in": [table]}} for table in tables_mentioned]
            if len(table_conditions) > 1:
                filter_conditions.append({"$or": table_conditions})
            else:
                filter_conditions.append(table_conditions[0])
        
        if figures_mentioned:
            # Filter for chunks that mention ANY of the specified figures
            figure_conditions = [{"figures_mentioned": {"$in": [figure]}} for figure in figures_mentioned]
            if len(figure_conditions) > 1:
                filter_conditions.append({"$or": figure_conditions})
            else:
                filter_conditions.append(figure_conditions[0])
        
        if clauses_mentioned:
            # Filter for chunks that mention ANY of the specified clauses
            clause_conditions = [{"clauses_mentioned": {"$in": [clause]}} for clause in clauses_mentioned]
            if len(clause_conditions) > 1:
                filter_conditions.append({"$or": clause_conditions})
            else:
                filter_conditions.append(clause_conditions[0])
        
        # Combine all filter conditions with AND
        metadata_filter = None
        if filter_conditions:
            if len(filter_conditions) == 1:
                metadata_filter = filter_conditions[0]
            else:
                metadata_filter = {"$and": filter_conditions}
        
        # Prepare search parameters
        search_params = {
            "namespace": namespace,
            "query": {
                "top_k": min(top_k, 100),  # Increase limit to get all chunks from a page
                # Pinecone embedding models do not accept an empty string.
                # If no query is supplied we pass a harmless placeholder string so the
                # request succeeds while the actual filtering is performed via metadata.
                "inputs": {"text": query.strip() if query and query.strip() else "metadata filter"}
            }
        }
        
        # Add filter if we have one
        if metadata_filter:
            search_params["query"]["filter"] = metadata_filter
        
        # No reranking - we want natural page order
        
        # Execute the search
        response = index.search(**search_params)
        
        # Process results
        hits = response.get("result", {}).get("hits", [])
        
        if not hits:
            filter_desc = []
            if source_document_id:
                filter_desc.append(f"document_id='{source_document_id}'")
            if page_number is not None:
                filter_desc.append(f"page={page_number}")
            if tables_mentioned:
                filter_desc.append(f"tables={tables_mentioned}")
            if figures_mentioned:
                filter_desc.append(f"figures={figures_mentioned}")
            if clauses_mentioned:
                filter_desc.append(f"clauses={clauses_mentioned}")
            
            return {
                "type": "filtered_search_results",
                "query": query,
                "filters_applied": " AND ".join(filter_desc) if filter_desc else "None",
                "message": f"No documents found matching the specified filters",
                "concatenated_content": "",
                "chunk_count": 0
            }
        
        # Collect all chunks with their metadata
        chunks = []
        for hit in hits:
            fields = hit.get("fields", {})
            chunks.append({
                "source_document_id": fields.get("source_document_id", "Unknown Document"),
                "page_number": fields.get("page_number", 0),
                "chunk_index_in_page": fields.get("chunk_index_in_page", 0),
                "content": fields.get("chunk_text", ""),
                "clauses_mentioned": fields.get("clauses_mentioned", []),
                "tables_mentioned": fields.get("tables_mentioned", []),
                "figures_mentioned": fields.get("figures_mentioned", [])
            })
        
        # Sort by page number and then by chunk index to maintain natural order
        chunks.sort(key=lambda x: (x.get("page_number", 0), x.get("chunk_index_in_page", 0)))
        
        # Concatenate all chunk text
        concatenated_content = "\n\n".join([chunk["content"] for chunk in chunks])
        
        # Build filter description for response
        filter_desc = []
        if source_document_id:
            filter_desc.append(f"document_id='{source_document_id}'")
        if page_number is not None:
            filter_desc.append(f"page={page_number}")
        if tables_mentioned:
            filter_desc.append(f"tables={tables_mentioned}")
        if figures_mentioned:
            filter_desc.append(f"figures={figures_mentioned}")
        if clauses_mentioned:
            filter_desc.append(f"clauses={clauses_mentioned}")
        
        # Collect unique mentions across all chunks
        all_clauses = set()
        all_tables = set()
        all_figures = set()
        for chunk in chunks:
            all_clauses.update(chunk.get("clauses_mentioned", []))
            all_tables.update(chunk.get("tables_mentioned", []))
            all_figures.update(chunk.get("figures_mentioned", []))
        
        return {
            "type": "filtered_search_results",
            "query": query if query else "No text query (metadata filtering only)",
            "filters_applied": " AND ".join(filter_desc) if filter_desc else "None",
            "concatenated_content": concatenated_content,
            "chunk_count": len(chunks),
            "pages_covered": sorted(set(chunk["page_number"] for chunk in chunks)),
            "all_clauses_mentioned": sorted(list(all_clauses)),
            "all_tables_mentioned": sorted(list(all_tables)),
            "all_figures_mentioned": sorted(list(all_figures)),
            "source_document": chunks[0]["source_document_id"] if chunks else "Unknown"
        }
        
    except Exception as e:
        return {
            "type": "database_search_error",
            "error": f"Error in filtered search: {str(e)}",
            "details": "Please check your filter parameters and try again."
        }

# Google Generative AI imports
try:
    import google.generativeai as genai
    from PIL import Image
    import asyncio
except ImportError:
    genai = None
    Image = None
    asyncio = None

# Excel integration imports
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    openpyxl = None
    Workbook = None
    load_workbook = None

# Excel Configuration
EXCEL_SPREADSHEETS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "company_spreadsheets")

@tool
def list_excel_spreadsheets(directory_path: str = None) -> Dict[str, Any]:
    """List Excel spreadsheets in the company spreadsheets directory with their metadata.

    Reads from the centralized ``spreadsheets_info.meta.json`` file that contains
    descriptions, inputs, outputs, and purpose for all Excel spreadsheets. This allows
    the LLM to understand how to interact with each spreadsheet.

    Args:
        directory_path: Custom directory to scan. Defaults to
            ``EXCEL_SPREADSHEETS_DIR``.

    Returns:
        A dictionary with a list of spreadsheets and their metadata from the meta.json file.
    """
    if openpyxl is None:
        return {
            "type": "error",
            "error": "openpyxl not installed. Please install with: pip install openpyxl",
        }

    try:
        target_dir = directory_path or EXCEL_SPREADSHEETS_DIR
        if not os.path.exists(target_dir):
            return {
                "type": "excel_files_list",
                "count": 0,
                "files": [],
                "message": f"Directory not found: {target_dir}",
            }

        # Load centralized metadata
        metadata_file = os.path.join(target_dir, "spreadsheets_info.meta.json")
        if not os.path.exists(metadata_file):
            return {
                "type": "excel_spreadsheets_list",
                "count": 0,
                "spreadsheets": [],
                "message": f"Metadata file not found: {metadata_file}",
            }

        try:
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata_content = json.load(f)
                excel_files = metadata_content.get("excel_files", [])
                
                # Verify that the Excel files actually exist
                verified_files = []
                for excel_info in excel_files:
                    file_path = os.path.join(target_dir, excel_info["filename"])
                    if os.path.exists(file_path):
                        # Add full path for tools to use
                        excel_info["full_path"] = file_path
                        verified_files.append(excel_info)
                
                return {
                    "type": "excel_spreadsheets_list",
                    "count": len(verified_files),
                    "spreadsheets": verified_files
                }
                
        except Exception as e:  # noqa: BLE001
            return {
                "type": "error",
                "error": f"Failed to parse spreadsheets_info.meta.json: {e}"
            }

    except Exception as e:  # noqa: BLE001
        return {"type": "error", "error": f"Error listing Excel files: {e}"}


@tool
def execute_excel_calculations(file_path: str, input_data: Dict[str, Union[str, int, float]], 
                             output_cells: List[str], sheet_name: str = None) -> Dict[str, Any]:
    """Execute a calculation in Excel by writing input values and reading output values.
    
    This function writes input values to specified cells, saves the file to trigger calculations,
    and then reads the results from output cells.
    
    Args:
        file_path: Path to Excel file (relative to company_spreadsheets directory)
        input_data: Dictionary mapping cell references to input values (e.g., {'A1': 10, 'B1': 20})
        output_cells: List of cell references to read results from (e.g., ['C1', 'D1'])
        sheet_name: Name of the worksheet (optional, uses first sheet if not specified)
        
    Returns:
        Dictionary containing input values, output values, and calculation results.
    """
    if openpyxl is None:
        return {
            "type": "error",
            "error": "openpyxl not installed. Please install with: pip install openpyxl"
        }
    
    debug_info = {"steps": [], "method_used": None}
    
    try:
        # Construct full file path
        if not os.path.isabs(file_path):
            full_path = os.path.join(EXCEL_SPREADSHEETS_DIR, file_path)
        else:
            full_path = file_path
        
        debug_info["steps"].append(f"Full file path: {full_path}")
        
        if not os.path.exists(full_path):
            return {
                "type": "error",
                "error": f"File not found: {full_path}"
            }
        
        # First, write the input values
        debug_info["steps"].append("Loading workbook for writing")
        wb = load_workbook(full_path)
        
        # Get worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return {
                    "type": "error",
                    "error": f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
                }
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title
        
        debug_info["steps"].append(f"Using sheet: {sheet_name}")
        
        # Write input values
        for cell_ref, value in input_data.items():
            ws[cell_ref] = value
            debug_info["steps"].append(f"Set cell {cell_ref} = {value}")
        
        # Save the file
        wb.save(full_path)
        wb.close()
        debug_info["steps"].append("Saved workbook with input values")
        
        # Try to use Excel COM automation if on Windows
        output_values = {}
        calculation_success = False
        
        try:
            # Try using pywin32 for Excel automation
            import win32com.client
            debug_info["method_used"] = "win32com"
            debug_info["steps"].append("Attempting Excel COM automation")
            
            # Create Excel application
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open workbook
            workbook = excel.Workbooks.Open(os.path.abspath(full_path))
            worksheet = workbook.Worksheets(sheet_name)
            
            # Force calculation
            excel.Calculate()
            debug_info["steps"].append("Forced Excel calculation")
            
            # Read output values
            for cell_ref in output_cells:
                cell_value = worksheet.Range(cell_ref).Value
                output_values[cell_ref] = cell_value
                debug_info["steps"].append(f"Read {cell_ref} = {cell_value}")
            
            # Close Excel
            workbook.Close(SaveChanges=True)
            excel.Quit()
            
            calculation_success = True
            debug_info["steps"].append("Excel COM automation successful")
            
        except ImportError:
            debug_info["steps"].append("win32com not available, trying alternative method")
        except Exception as e:
            debug_info["steps"].append(f"Excel COM failed: {str(e)}")
        
        # If COM automation failed, try openpyxl with formulas
        if not calculation_success:
            debug_info["method_used"] = "openpyxl_formula_evaluation"
            debug_info["steps"].append("Attempting formula evaluation with openpyxl")
            
            # Re-open with openpyxl to read formulas and try to evaluate them
            wb = load_workbook(full_path, data_only=False)
            ws = wb[sheet_name]
            
            # Simple formula evaluator for basic calculations
            def evaluate_simple_formula(formula: str, worksheet) -> Union[float, str]:
                """Evaluate simple Excel formulas"""
                if not formula or not isinstance(formula, str) or not formula.startswith('='):
                    return formula
                
                # Remove the '=' sign
                formula = formula[1:]
                
                # Replace cell references with their values
                import re
                cell_pattern = r'([A-Z]+)(\d+)'
                
                def replace_cell_ref(match):
                    col = match.group(1)
                    row = match.group(2)
                    cell_ref = f"{col}{row}"
                    cell_value = worksheet[cell_ref].value
                    
                    # If the cell contains a formula, try to get its calculated value
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        # For now, return 0 for formula cells
                        return "0"
                    
                    return str(cell_value) if cell_value is not None else "0"
                
                # Replace cell references with values
                formula_with_values = re.sub(cell_pattern, replace_cell_ref, formula)
                
                try:
                    # Simple evaluation for basic arithmetic
                    # This is limited but safe for basic calculations
                    result = eval(formula_with_values, {"__builtins__": {}}, {})
                    return float(result)
                except:
                    return f"FORMULA: {formula}"
            
            # Read output cells and try to evaluate formulas
            for cell_ref in output_cells:
                cell = ws[cell_ref]
                cell_value = cell.value
                
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    # Try to evaluate the formula
                    evaluated = evaluate_simple_formula(cell_value, ws)
                    output_values[cell_ref] = evaluated
                    debug_info["steps"].append(f"Evaluated {cell_ref}: {cell_value} = {evaluated}")
                else:
                    output_values[cell_ref] = cell_value
                    debug_info["steps"].append(f"Read {cell_ref} = {cell_value}")
            
            wb.close()
            
            # If we still don't have values, try reading with data_only
            if all(isinstance(v, str) and v.startswith('FORMULA:') for v in output_values.values()):
                debug_info["steps"].append("Formula evaluation incomplete, trying data_only mode")
                wb = load_workbook(full_path, data_only=True)
                ws = wb[sheet_name]
                
                temp_values = {}
                for cell_ref in output_cells:
                    temp_values[cell_ref] = ws[cell_ref].value
                
                wb.close()
                
                # Use data_only values if they're not None
                for cell_ref, value in temp_values.items():
                    if value is not None:
                        output_values[cell_ref] = value
                        debug_info["steps"].append(f"Got cached value for {cell_ref} = {value}")
        
        return {
            "type": "excel_calculation_success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "inputs": input_data,
            "outputs": output_values,
            "debug_info": debug_info,
            "message": "Calculation completed"
        }
        
    except Exception as e:
        return {
            "type": "error",
            "error": f"Error executing Excel calculation: {str(e)}",
            "debug_info": debug_info
        }


@tool
async def analyze_document_vision(
    document_id: str,
    page_number: int,
    query: str
) -> Dict[str, Any]:
    """Analyze a specific page image using Google's Gemini 2.5 Pro vision model.
    
    This tool loads a pre-processed page image from the page_images directory
    and uses Gemini's vision capabilities to analyze it based on the provided query.
    Useful for extracting information from diagrams, charts, tables, or complex
    layouts that require visual understanding.
    
    Args:
        document_id: Document identifier (e.g., 'as_4100_1998' or 'as_1170.0_2002')
        page_number: Page number to analyze (1-indexed)
        query: Specific question or analysis request for the page
    
    Returns:
        A dictionary containing the vision analysis results or error information
    """
    if genai is None or Image is None:
        return {
            "type": "error",
            "error": "Required libraries not installed. Please install: pip install google-generativeai pillow"
        }
    
    # Check for API key
    GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
    if not GOOGLE_API_KEY:
        return {
            "type": "error",
            "error": "GOOGLE_API_KEY environment variable not set"
        }
    
    try:
        # Configure Gemini
        genai.configure(api_key=GOOGLE_API_KEY)
        
        # Initialize the model - using gemini-2.5-pro which has vision capabilities
        model = genai.GenerativeModel('gemini-2.5-pro')
        
        # Construct the image path
        # Images are stored in subdirectories: base_dir/document_id/page_X.png
        base_image_dir = r"C:\Users\gabri\Desktop\Engineering\aust_standards_digitilization\agentic_rag_v2\agent-chat-ui\public\data\page_images"
        
        # Build the path to the specific page image
        image_filename = f"page_{page_number}.png"
        image_path = os.path.join(base_image_dir, document_id, image_filename)
        
        # Check if the image exists
        if not os.path.exists(image_path):
            # Check if the document directory exists
            doc_dir = os.path.join(base_image_dir, document_id)
            if not os.path.exists(doc_dir):
                # List available document directories
                available_docs = []
                if os.path.exists(base_image_dir):
                    available_docs = [d for d in os.listdir(base_image_dir) 
                                    if os.path.isdir(os.path.join(base_image_dir, d))]
                
                return {
                    "type": "error",
                    "error": f"Document directory '{document_id}' not found",
                    "available_documents": available_docs,
                    "base_directory": base_image_dir
                }
            else:
                # List available pages for this document
                available_pages = []
                for file in os.listdir(doc_dir):
                    if file.startswith("page_") and file.endswith(".png"):
                        try:
                            page_num = int(file.replace("page_", "").replace(".png", ""))
                            available_pages.append(page_num)
                        except:
                            pass
                
                available_pages.sort()
                
                return {
                    "type": "error",
                    "error": f"Page {page_number} not found for document '{document_id}'",
                    "available_pages": available_pages,
                    "document_directory": doc_dir
                }
        
        # Load the image using asyncio.to_thread to avoid blocking
        img = await asyncio.to_thread(Image.open, image_path)
        
        # Get image metadata in a non-blocking way
        img_width = img.width
        img_height = img.height
        img_format = img.format
        img_mode = img.mode
        
        # Prepare the prompt
        prompt = f"""Please analyze this page image from a document and answer the following query:

Query: {query}

Additional context - this is page {page_number} of document '{document_id}'.
Please provide a detailed and accurate response based on what you can see in the image.
Focus on:
- Text content (including headers, paragraphs, and footnotes)
- Diagrams, charts, or technical drawings
- Tables and structured data
- Mathematical equations or formulas
- Any visual elements relevant to the query

Be specific and reference exact content from the image when answering."""

        # Generate response using vision model (this is already async)
        response = await asyncio.to_thread(model.generate_content, [prompt, img])
        
        # Extract the response text
        if response.text:
            analysis_result = response.text
        else:
            analysis_result = "No analysis could be generated from the image."
        
        return {
            "type": "vision_analysis_success",
            "document_id": document_id,
            "page_number": page_number,
            "query": query,
            "analysis": analysis_result,
            "image_path": image_path,
            "metadata": {
                "image_size": f"{img_width} x {img_height}",
                "image_format": img_format,
                "image_mode": img_mode
            }
        }
        
    except Exception as e:
        return {
            "type": "error",
            "error": f"Error during vision analysis: {str(e)}",
            "document_id": document_id,
            "page_number": page_number
        }


# Add your tools to this list. The ReAct agent will be able to invoke them.

@tool
async def get_document_page_text(
    document_id: str,
    page_number: int,
) -> Dict[str, Any]:
    """Retrieve markdown text for a specific page of a document.

    The preprocessing pipeline saves OCR-extracted text for every page as a
    ``.md`` file in the following structure::

        <base_dir>/<document_id>/<page_number>.md

    Example::
        as_1170.0_2002/1.md

    Args:
        document_id: Identifier of the source document (e.g., ``"as_1170.0_2002"``).
        page_number: 1-indexed page number.

    Returns:
        A dictionary with either the page text (``type = 'page_text_success'``)
        or helpful error information if the file cannot be found.
    """
    import asyncio  # local import avoids global dependency if asyncio missing

    base_text_dir = (
        r"C:\\Users\\gabri\\Desktop\\Engineering\\aust_standards_digitilization\\agentic_rag_v2\\agent-chat-ui\\public\\data\\page_text"
    )

    try:
        # Build expected file path
        file_path = os.path.join(base_text_dir, document_id, f"{page_number}.md")

        # Handle missing file or directory with informative errors
        if not os.path.exists(file_path):
            doc_dir = os.path.join(base_text_dir, document_id)
            if not os.path.isdir(doc_dir):
                available_docs = []
                if os.path.exists(base_text_dir):
                    available_docs = [d for d in os.listdir(base_text_dir) if os.path.isdir(os.path.join(base_text_dir, d))]
                return {
                    "type": "error",
                    "error": f"Document directory '{document_id}' not found",
                    "available_documents": available_docs,
                    "base_directory": base_text_dir,
                }

            # Directory exists, so list available page numbers for helpful feedback
            available_pages: list[int] = []
            for fname in os.listdir(doc_dir):
                if fname.endswith(".md"):
                    try:
                        available_pages.append(int(fname.replace(".md", "")))
                    except ValueError:
                        pass
            available_pages.sort()
            return {
                "type": "error",
                "error": f"Page {page_number} not found for document '{document_id}'",
                "available_pages": available_pages,
                "document_directory": doc_dir,
            }

        # Read markdown content in a separate thread to avoid blocking
        page_text = await asyncio.to_thread(lambda: Path(file_path).read_text(encoding="utf-8"))

        return {
            "type": "page_text_success",
            "document_id": document_id,
            "page_number": page_number,
            "text": page_text,
            "length": len(page_text),
            "path": file_path,
        }

    except Exception as e:  # noqa: BLE001
        return {
            "type": "error",
            "error": f"Error reading page text: {e}",
            "document_id": document_id,
            "page_number": page_number,
        }
# Import Google Sheets tools
from .sheets_tool import sheets_calculate, list_sheets_calculators

# The tools should be functions that accept a single string argument and return a string.
# The docstrings of the functions will be used to tell the LLM about the tool.
TOOLS: List[Callable[[Any], Any]] = [
    search, 
    calculator, 
    search_engineering_database, 
    # search_engineering_database_filtered,
    list_excel_spreadsheets,
    execute_excel_calculations,
    sheets_calculate,  # Google Sheets calculator
    list_sheets_calculators,  # List available Google Sheets
    analyze_document_vision,
    get_document_page_text
]
