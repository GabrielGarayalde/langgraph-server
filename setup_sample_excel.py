#!/usr/bin/env python3
"""
Script to create sample Excel calculation files for testing the Excel integration tools.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_beam_calculation_sheet():
    """Create a sample beam calculation Excel file."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Beam Calculations"
    
    # Headers and styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    output_fill = PatternFill(start_color="E7FFE7", end_color="E7FFE7", fill_type="solid")
    
    # Title
    ws["A1"] = "STRUCTURAL BEAM CALCULATIONS"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    
    # Input section
    ws["A3"] = "INPUTS"
    ws["A3"].font = header_font
    ws["A3"].fill = header_fill
    ws.merge_cells("A3:B3")
    
    ws["A4"] = "Load (kN)"
    ws["B4"] = 10  # Default value
    ws["A4"].fill = input_fill
    ws["B4"].fill = input_fill
    
    ws["A5"] = "Span (m)"
    ws["B5"] = 6  # Default value
    ws["A5"].fill = input_fill
    ws["B5"].fill = input_fill
    
    ws["A6"] = "Section Modulus (cm³)"
    ws["B6"] = 500  # Default value
    ws["A6"].fill = input_fill
    ws["B6"].fill = input_fill
    
    # Calculation section
    ws["A8"] = "CALCULATIONS"
    ws["A8"].font = header_font
    ws["A8"].fill = header_fill
    ws.merge_cells("A8:D8")
    
    ws["A9"] = "Maximum Moment (kNm)"
    ws["B9"] = "=B4*B5^2/8"  # Formula: wL²/8 for uniformly distributed load
    ws["A9"].fill = output_fill
    ws["B9"].fill = output_fill
    
    ws["A10"] = "Maximum Stress (MPa)"
    ws["B10"] = "=B9*1000000/(B6*1000)"  # Formula: M/S (convert units)
    ws["A10"].fill = output_fill
    ws["B10"].fill = output_fill
    
    ws["A11"] = "Deflection Factor"
    ws["B11"] = "=5*B4*B5^4/(384*200000*B6*10000)"  # Simplified deflection calculation
    ws["A11"].fill = output_fill
    ws["B11"].fill = output_fill
    
    # Results section
    ws["A13"] = "RESULTS"
    ws["A13"].font = header_font
    ws["A13"].fill = header_fill
    ws.merge_cells("A13:D13")
    
    ws["A14"] = "Status"
    ws["B14"] = '=IF(B10<=250,"PASS","FAIL")'  # Check if stress < 250 MPa
    ws["A14"].fill = output_fill
    ws["B14"].fill = output_fill
    
    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    
    # Save file
    file_path = "company_spreadsheets/structural/beam_calculations.xlsx"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    print(f"Created: {file_path}")

def create_simple_calculator():
    """Create a simple calculator Excel file."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculator"
    
    # Simple calculator layout
    ws["A1"] = "SIMPLE CALCULATOR"
    ws["A1"].font = Font(bold=True, size=14)
    
    ws["A3"] = "Number 1"
    ws["B3"] = 0
    
    ws["A4"] = "Number 2"
    ws["B4"] = 0
    
    ws["A6"] = "Sum"
    ws["B6"] = "=B3+B4"
    
    ws["A7"] = "Product"
    ws["B7"] = "=B3*B4"
    
    ws["A8"] = "Difference"
    ws["B8"] = "=B3-B4"
    
    ws["A9"] = "Division"
    ws["B9"] = "=IF(B4<>0,B3/B4,\"Error: Division by zero\")"
    
    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    
    # Save file
    file_path = "company_spreadsheets/calculator.xlsx"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    print(f"Created: {file_path}")

if __name__ == "__main__":
    print("Creating sample Excel calculation files...")
    create_beam_calculation_sheet()
    create_simple_calculator()
    print("Sample files created successfully!")
